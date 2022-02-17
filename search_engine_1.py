#! python3
# wyszukiwarka_1.py

import requests, sys, webbrowser, bs4
from bs4 import BeautifulSoup
import re
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import date

today = date.today()
d1 = today.strftime("%d-%m-%Y")

wb = openpyxl.load_workbook('Your_assets_up_to_date.xlsx')
prices_sheet = wb.get_sheet_by_name('PRICES')


sites_1 =['https://www.bankier.pl/waluty/kursy-walut/forex/USDPLN',
        'https://www.bankier.pl/waluty/kursy-walut/forex/EURPLN',
        'https://www.bankier.pl/waluty/kursy-walut/forex/CHFPLN',
        'https://www.bankier.pl/inwestowanie/profile/quote.html?symbol=PZU',
        'https://www.bankier.pl/inwestowanie/profile/quote.html?symbol=PKNORLEN',
        'https://www.bankier.pl/inwestowanie/profile/quote.html?symbol=ECHO',
        'https://www.bankier.pl/inwestowanie/profile/quote.html?symbol=ZLOTO',
        'https://www.bankier.pl/inwestowanie/profile/quote.html?symbol=SREBRO'
        ]

sites_2 =['https://coinmarketcap.com/currencies/bitcoin/',
        'https://coinmarketcap.com/currencies/ethereum/',
        'https://coinmarketcap.com/currencies/bnb/',
        'https://coinmarketcap.com/currencies/xrp/',
        'https://coinmarketcap.com/currencies/polkadot-new/',
        'https://coinmarketcap.com/currencies/solana/',
        'https://coinmarketcap.com/currencies/tron/'
        ]




for i in range(8):


    res = requests.get(sites_1[i])
    res.raise_for_status()

    soup = bs4.BeautifulSoup(res.text, 'html.parser')

    result = soup.select('.profilLast')

    number_v0 = result[0].getText()

    number_v1 = re.findall("\d+\,\d+", str(number_v0))
    number_v2 = number_v1[0]
    number_v3 = number_v2.replace(",", ".")
    number_v4 = float(number_v3)

    prices_sheet[str(get_column_letter(i+3))+'4'] = number_v4




for i in range(7):


    res_crypto = requests.get(sites_2[i])
    res_crypto.raise_for_status()

    soup_crypto = bs4.BeautifulSoup(res_crypto.text, 'html.parser')

    result_crypto = soup_crypto.select('.priceValue')

    number_v0_crypto = result_crypto[0].getText()


    number_v1_crypto = number_v0_crypto.replace("$", "")
    number_v2_crypto = number_v1_crypto.replace(",","")
    number_v3 = number_v2_crypto.replace(",", ".")
    number_v4 = float(number_v3)

    prices_sheet[str(get_column_letter(i+11))+'4'] = number_v4


sheet = wb.get_sheet_by_name('CURRENT STATE')
sheet['A5'] = 'Today:'
sheet['A6'] = d1





wb.save('Your_assets_up_to_date_'+str(d1)+'.xlsx')






